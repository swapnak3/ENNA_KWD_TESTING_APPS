# -*- coding: utf-8 -*-
"""Created on 27.02.2024.

@project: .
@author: SPLATZP: Pascal Platzer.

Contains helper functions for xpath.
"""
import logging
import pathlib
import time

import openpyxl as xl

from enna_kwd_testing.definitions import RESOURCES_PATH

MODULE_LOGGER = logging.getLogger(__name__)


# pylint: disable=protected-access, disable=too-many-try-statements

def get_warning_notification_data(warn_id) -> dict:
	"""Get warning notification data from xlsx worksheet.

	:param str warn_id: The ID rof the warning notification e.g. "0xA37D".

	:return: The generated warning notification data.
	:rtype: dict.
	"""

	warning_data = {}
	classification = ["WarningStore", "PictogramRoll", "WarningLights", "OnlineCarCare", "ServiceKey", "BreakdownCall", "HistoryStore"]
	path = "warning_notifications/DWM_WDB_AU416_V020_BL3.20.xlsx"

	worksheet__warning_data = get_xlsx_data(path)
	last_element = list(worksheet__warning_data._cells.keys())[-1][0]
	first_element = 1
	warn_id_time = int(time.time())

	try:
		for row_index in range(first_element, last_element):
			if warn_id in worksheet__warning_data.cell(row_index + 1, 1).value:
				if worksheet__warning_data.cell(row_index + 1, 35).value == "unavailable":
					continue
				warn_id = int(worksheet__warning_data.cell(row_index + 1, 1).value, 16)
				warn_id_hex = worksheet__warning_data.cell(row_index + 1, 1).value
				priority = int(worksheet__warning_data.cell(row_index + 1, 8).value)
				if worksheet__warning_data.cell(row_index + 1, 5).value is None:
					icon_color = "red"
				else:
					icon_color = worksheet__warning_data.cell(row_index + 1, 5).value

				try:
					icon_id = int(worksheet__warning_data.cell(row_index + 1, 4).value, 16)
				except ValueError as exception:
					MODULE_LOGGER.error(f"{exception}. No correct icon ID")
					icon_id = 0

				icon_id_hex = worksheet__warning_data.cell(row_index + 1, 4).value

				warning_data.update(
					{
						"classification": classification,
						"notificationId": warn_id, "warn_id": warn_id_hex,
						"presentationState": "finished",
						"priority": priority, "shownAt": warn_id_time,
						"userConfirmationRequired": "isOptional", "iconColor": icon_color,
						"iconID": icon_id, "iconID_hex": icon_id_hex
					}
				)
				break

		return warning_data
	except ValueError as exception:
		MODULE_LOGGER.error(f"{exception}.")
		return warning_data


def get_xlsx_data(source_path):
	"""Generate XPath based on the element type.

		:param str source_path: path to source.

		:return: The generated data.
		:rtype: bool | worksheet.
		"""
	try:
		path = pathlib.Path(".").absolute()
		path = path.joinpath(f"{RESOURCES_PATH}/{source_path}")

		wb = xl.load_workbook(path, data_only=True)
		data = wb.worksheets[0]
		wb.close()
	except FileNotFoundError as exception:
		MODULE_LOGGER.error(f"{exception}.")
		return False

	return data
